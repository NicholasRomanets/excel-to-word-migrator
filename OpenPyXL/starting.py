from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

wb = Workbook()
wb = load_workbook(filename='../Source.xlsx')
wb = wb.active
print(wb['F2'].value)